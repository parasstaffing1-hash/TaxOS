from __future__ import annotations

import io
from typing import Any, Optional
from decimal import Decimal

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import BarChart, Reference
except ImportError:
    openpyxl = None


class ExcelDocumentGenerator:
    """Generates professional Excel reports from calculator results using openpyxl."""

    def generate(self, calculator_config: Any, results: dict[str, Any], template: Any, inputs_data: Optional[dict[str, Any]] = None) -> bytes:
        """
        Generate an Excel report.

        Args:
            calculator_config: Configuration of the calculator.
            results: Calculated results dictionary.
            template: Report template configuration.
            inputs_data: Original user inputs.

        Returns:
            Excel file content as bytes.
        """
        if openpyxl is None:
            raise RuntimeError("openpyxl is not installed. Please install it to generate Excel files.")
            
        inputs_data = inputs_data or {}
        
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        if ws_summary is not None:
            ws_summary.title = "Summary"
        
        ws_inputs = wb.create_sheet("Inputs")
        ws_breakdown = wb.create_sheet("Breakdown")
        ws_charts = wb.create_sheet("Charts")
        
        primary_color_hex = getattr(getattr(template, 'branding', None), 'primary_color', '#000000')
        hex_color = primary_color_hex.lstrip('#')
        if len(hex_color) != 6:
            hex_color = "000000"
            
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        
        currency_symbol = getattr(getattr(template, 'locale', None), 'currency_symbol', '$')
        
        def apply_header_style(ws, row_idx):
            for cell in ws[row_idx]:
                cell.font = header_font
                cell.fill = header_fill
                
        def format_cell(cell, format_type):
            if format_type == 'currency':
                # Built-in currency format or custom
                cell.number_format = '"$"#,##0.00'
            elif format_type == 'percentage':
                cell.number_format = '0.00%'
                
        # --- Summary Worksheet ---
        title = getattr(calculator_config, 'title', 'Report')
        ws_summary.append([title])
        ws_summary.append(["Label", "Value"])
        apply_header_style(ws_summary, 2)
        ws_summary.freeze_panes = "A3"
        
        formulas = getattr(calculator_config, 'formulas', [])
        result_formulas = [f for f in formulas if getattr(f, 'is_result', False)]
        for f in result_formulas:
            f_id = getattr(f, 'id', '')
            f_label = getattr(f, 'label', f_id)
            f_format = getattr(f, 'format', 'default')
            val = results.get(f_id)
            if isinstance(val, Decimal):
                val = float(val)
            ws_summary.append([f_label, val])
            format_cell(ws_summary.cell(row=ws_summary.max_row, column=2), f_format)
            
        # --- Inputs Worksheet ---
        ws_inputs.append(["Input", "Value"])
        apply_header_style(ws_inputs, 1)
        ws_inputs.freeze_panes = "A2"
        ws_inputs.auto_filter.ref = ws_inputs.dimensions
        
        inputs = getattr(calculator_config, 'inputs', [])
        for inp in inputs:
            inp_id = getattr(inp, 'id', '')
            inp_label = getattr(inp, 'label', inp_id)
            val = inputs_data.get(inp_id, '')
            if isinstance(val, Decimal):
                val = float(val)
            ws_inputs.append([inp_label, val])
            
        # --- Breakdown Worksheet ---
        ws_breakdown.append(["ID", "Label", "Value", "Format"])
        apply_header_style(ws_breakdown, 1)
        ws_breakdown.freeze_panes = "A2"
        ws_breakdown.auto_filter.ref = ws_breakdown.dimensions
        
        for idx, f in enumerate(formulas, start=2):
            f_id = getattr(f, 'id', '')
            f_label = getattr(f, 'label', f_id)
            f_format = getattr(f, 'format', 'default')
            val = results.get(f_id)
            if isinstance(val, Decimal):
                val = float(val)
            ws_breakdown.append([f_id, f_label, val, f_format])
            format_cell(ws_breakdown.cell(row=ws_breakdown.max_row, column=3), f_format)
            
        # --- Charts Worksheet ---
        chart = BarChart()
        chart.title = "Results Breakdown"
        chart.x_axis.title = "Item"
        chart.y_axis.title = "Value"
        
        if len(formulas) > 0:
            data = Reference(ws_breakdown, min_col=3, min_row=1, max_row=len(formulas)+1)
            cats = Reference(ws_breakdown, min_col=2, min_row=2, max_row=len(formulas)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws_charts.add_chart(chart, "A1")
            
        # Auto column width
        for ws in [ws_summary, ws_inputs, ws_breakdown]:
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2
                
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
