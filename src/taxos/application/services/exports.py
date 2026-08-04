"""Generators for PDF and Excel exports."""

from __future__ import annotations

import csv
import io
from decimal import Decimal

from fpdf import FPDF
from openpyxl import Workbook

from taxos.api.schemas.calculator import CalculationResponse


class PDFReport(FPDF):
    """Custom PDF class for TaxOS reports."""

    def header(self) -> None:
        self.set_font("helvetica", "B", 15)
        self.cell(0, 10, "TaxOS After-Tax Salary Report", border=0, align="C")
        self.ln(20)

    def footer(self) -> None:
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")


def generate_pdf_report(result: CalculationResponse) -> bytes:
    """Generate a PDF byte string from the CalculationResponse."""
    pdf = PDFReport()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 10, "Summary", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("helvetica", "", 10)

    # Currency symbol (basic support)
    sym = "$" if result.currency.value == "USD" else result.currency.value

    def add_row(label: str, amount: Decimal) -> None:
        pdf.cell(80, 8, label)
        pdf.cell(40, 8, f"{sym}{amount:,.2f}", new_x="LMARGIN", new_y="NEXT", align="R")

    add_row("Gross Annual Income:", result.gross_income.annual)
    add_row("Net Annual Income:", result.net_income.annual)
    add_row("Effective Tax Rate:", result.effective_tax_rate)
    add_row("Marginal Tax Rate:", result.marginal_tax_rate or Decimal("0"))

    pdf.ln(10)
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 10, "Tax Breakdown", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("helvetica", "", 10)
    for item in result.breakdown:
        pdf.cell(80, 8, item.name or item.rule)
        pdf.cell(40, 8, f"{sym}{item.tax:,.2f}", new_x="LMARGIN", new_y="NEXT", align="R")

    # Return as bytes
    return bytes(pdf.output())


def generate_excel_report(result: CalculationResponse) -> bytes:
    """Generate an Excel byte string from the CalculationResponse."""
    wb = Workbook()
    ws = wb.active
    if not ws:
        raise RuntimeError("No active worksheet")
    ws.title = "Tax Report"

    # Headers
    ws.append(["Category", "Amount", "Details"])

    sym = result.currency.value

    # Summary
    ws.append(["Gross Annual Income", float(result.gross_income.annual), sym])
    ws.append(["Net Annual Income", float(result.net_income.annual), sym])
    ws.append(["Total Tax", float(result.total_tax or Decimal("0")), sym])
    ws.append(["Effective Tax Rate", float(result.effective_tax_rate), "%"])
    ws.append(["Marginal Tax Rate", float(result.marginal_tax_rate or 0), "%"])

    ws.append([])
    ws.append(["Detailed Breakdown", "Tax", "Deduction", "Credit", "Employer Cost"])

    for item in result.breakdown:
        ws.append(
            [
                item.name or item.rule,
                float(item.tax),
                float(item.deduction),
                float(item.credit),
                float(item.employer_cost),
            ]
        )

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def generate_csv_report(result: CalculationResponse) -> bytes:
    """Generate a CSV byte string from the CalculationResponse."""
    stream = io.StringIO()
    writer = csv.writer(stream)

    sym = result.currency.value

    # Summary Section
    writer.writerow(["Category", "Amount", "Details"])
    writer.writerow(["Gross Annual Income", str(result.gross_income.annual), sym])
    writer.writerow(["Net Annual Income", str(result.net_income.annual), sym])
    writer.writerow(["Total Tax", str(result.total_tax), sym])
    writer.writerow(["Employer Cost", str(result.employer_cost.annual), sym])
    writer.writerow(["Employee Deductions", str(result.employee_deductions.annual), sym])
    writer.writerow(["Effective Tax Rate", str(result.effective_tax_rate), "%"])
    writer.writerow(["Marginal Tax Rate", str(result.marginal_tax_rate or 0), "%"])

    writer.writerow([])

    # Detailed Breakdown Section
    writer.writerow(["Detailed Breakdown", "Tax", "Deduction", "Credit", "Employer Cost"])
    for item in result.breakdown:
        writer.writerow(
            [
                item.name or item.rule,
                str(item.tax),
                str(item.deduction),
                str(item.credit),
                str(item.employer_cost),
            ]
        )

    return stream.getvalue().encode("utf-8")
